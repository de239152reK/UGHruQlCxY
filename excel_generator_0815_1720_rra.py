# 代码生成时间: 2025-08-15 17:20:42
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException
"""
Excel表格自动生成器
使用Python和Numpy框架生成Excel表格文件
"""

class ExcelGenerator:
    """Excel表格生成器类"""
    
    def __init__(self, filename, sheet_name='Sheet1'):
        """初始化函数"""
        self.filename = filename
        self.sheet_name = sheet_name
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = sheet_name

    def generate_header(self, header_list):
        """生成表头"""
        if not isinstance(header_list, list):
            raise ValueError("表头必须是列表类型")
        for col, header in enumerate(header_list, start=1):
            self.ws.cell(row=1, column=col, value=header)

    def generate_data(self, data_list):
        """生成数据行"""
        if not isinstance(data_list, list):
            raise ValueError("数据行必须是列表类型")
        for row, row_data in enumerate(data_list, start=2):
            if not isinstance(row_data, list):
                raise ValueError("每行数据必须是列表类型")
            for col, data in enumerate(row_data, start=1):
                self.ws.cell(row=row, column=col, value=data)

    def save_excel(self):
        """保存Excel文件"""
        try:
            self.wb.save(self.filename)
            print(f"Excel文件{self.filename}已成功生成")
        except IOError as e:
            print(f"保存Excel文件失败：{e}")
        except InvalidFileException as e:
            print(f"无效的文件名：{e}")

    def generate_excel(self, header_list, data_list):
        """生成Excel文件"""
        self.generate_header(header_list)
        self.generate_data(data_list)
        self.save_excel()

# 示例用法
if __name__ == '__main__':
    header_list = ['姓名', '年龄', '性别']
    data_list = [['张三', 28, '男'], ['李四', 25, '女']]
    generator = ExcelGenerator('example.xlsx')
    generator.generate_excel(header_list, data_list)